import tkinter as tk #for GUI
from tkinter import filedialog #fungsi untuk memilih file atau folder
import os #operation system
from PIL import Image #membuka, memanipulasi, dan save berbagai format gambar
from datetime import datetime #tanggal dan waktu
from openpyxl import Workbook, load_workbook #menulis dan membaca excel
from openpyxl.styles import Border, Side, PatternFill

class HexEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Simple Hex Editor")
        self.input_files = []
        self.output_folder = None
        self.input_folder_test_code = []
        self.output_folder_test_code = None
        self.create_widgets()

    def create_widgets(self):
        self.text_widget = tk.Text(self.root, wrap="none", undo=True, width=80, height=20)
        self.text_widget.pack(expand=True, fill="both")

        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open", command=self.open_file)
        file_menu.add_command(label="Save", command=self.save_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.destroy)

        file_menu_hex_auto = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Hex Auto", menu=file_menu_hex_auto)
        file_menu_hex_auto.add_command(label="Select File Input", command=self.open_file_hex_auto)
        file_menu_hex_auto.add_command(label="Select Output Folder", command=self.select_output_folder)
        file_menu_hex_auto.add_command(label="Carve Files", command=self.carve_files)

        file_menu_test_code = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Test Code", menu=file_menu_test_code)
        file_menu_test_code.add_command(label="Select Input Folder", command=self.select_input_folder_test_code)
        file_menu_test_code.add_command(label="Select Output Folder", command=self.select_output_folder_test_code)
        file_menu_test_code.add_command(label="Execute", command=self.execute_test_code)
    #hex editor
    def open_file(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            with open(file_path, "rb") as file:
                content = file.read()
                hex_lines = [content[i:i+10] for i in range(0, len(content), 10)]
                ascii_lines = [self.format_ascii_line(content[i:i+10]) for i in range(0, len(content), 10)]

                formatted_content = ""
                for hex_line, ascii_line in zip(hex_lines, ascii_lines):
                    formatted_content += f"{self.format_hex_line(hex_line)}  |  {ascii_line}\n"

                self.text_widget.delete(1.0, tk.END)
                self.text_widget.insert(1.0, f"Offset   |  Hexadecimal                           |  ASCII\n")
                self.text_widget.insert(2.0, f"--------+--------------------------------------+------------------------\n")
                self.text_widget.insert(3.0, formatted_content)

    def format_hex_line(self, hex_line):
        return " ".join(f"{byte:02X}" for byte in hex_line)

    def format_ascii_line(self, ascii_line):
        return "".join(chr(ascii_char) if 32 <= ascii_char < 127 else '.' for ascii_char in ascii_line)

    def save_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".bin", filetypes=[("Binary Files", "*.bin")])
        if file_path:
            hex_content = self.text_widget.get(3.0, tk.END)
            hex_values = [int(value, 16) for value in hex_content.split() if value]
            byte_values = bytes(hex_values)
            with open(file_path, "wb") as file:
                file.write(byte_values)
#fitur hex auto
    def open_file_hex_auto(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("All Files", "*.*")])
        for file_path in file_paths:
           self.input_files.append(file_path)
        self.text_widget.insert(tk.END, 'Selected Input Files: \n')
        self.text_widget.insert(tk.END, '\n'.join(self.input_files))
        self.text_widget.insert(tk.END, '\n\n')

    def select_output_folder(self):
        self.output_folder = filedialog.askdirectory(title="Select Output Folder")
        if self.output_folder:
            self.text_widget.insert(tk.END, "Selected Output Folder: " + self.output_folder + "\n")

    def carve_files(self):
        if not self.input_files:
            print("Please select input files first.")
            return

        if not self.output_folder:
            print("Please select an output folder first.")
            return        
        
        combined_data = bytearray()
        file_counter = 1  # Untuk memberikan nama unik kepada setiap file

        # Iterate over the input files
        for input_file in self.input_files:
            with open(input_file, 'rb') as file:
                # Read the entire content of the file and append to the combined data
                combined_data += file.read()

        file_info = {
            'jpg': {
                'signature': b'\xFF\xD8\xFF\xE0',
                'end_marker': b'\xFF\xD9'
            },
            'png': {
                'signature': b'\x89\x50\x4E\x47',
                'end_marker': b'IEND\xAE\x42\x60\x82'
            },
            'gif': {
                'signature': b'\x47\x49\x46\x38',
                'end_marker': b'\x3B'
            },
            # Add more file types as needed
        }

        # Iterate over the file signatures
        for extension, info in file_info.items():
            start_index = combined_data.find(info['signature'])

            # Continue searching until the signature is not found
            while start_index != -1:
                end_index = combined_data.find(info['end_marker'], start_index)

                # Check if the ending index is found
                if end_index != -1:
                    # Carve the file data between the start and end indices
                    file_data = combined_data[start_index:end_index + len(info['end_marker'])]

                    # Create the output file path with a unique name
                    output_file = os.path.join(self.output_folder, f'carved_file_{file_counter:03d}.{extension}')

                    # Write the carved file data to the output file
                    with open(output_file, 'wb') as output_file:
                        output_file.write(file_data)

                    # Print information about the extracted file
                    print(f'File {output_file} extracted successfully.')

                    # Update the start index to continue searching
                    start_index = combined_data.find(info['signature'], end_index + len(info['end_marker']))

                    # Increment the file counter for the next unique name
                    file_counter += 1
                else:
                    # Signature is found but ending index is not found, break the loop
                    break
#fitur import excel
    def select_input_folder_test_code(self):
        while True:
            folder = filedialog.askdirectory(title="Select Input Folder")
            
            if not folder:
                break
            
            self.input_folder_test_code.append(folder)
        
        if self.input_folder_test_code:
            self.text_widget.insert(tk.END, 'Selected Input Folders: \n')
            self.text_widget.insert(tk.END, '\n'.join(self.input_folder_test_code))
            self.text_widget.insert(tk.END, '\n\n')
        else:
            self.text_widget.insert(tk.END, "No input folders selected\n")

    def select_output_folder_test_code(self):
        self.output_folder_test_code = filedialog.askdirectory(title="Select Output Folder")
        if self.output_folder_test_code:
            self.text_widget.insert(tk.END, "Selected Output Folder: " + self.output_folder_test_code + "\n")

    def extract_image_info(self, file_path):
        try:
            with Image.open(file_path) as img:
                created_time = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')
                modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')
                accessed_time = datetime.fromtimestamp(os.path.getatime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')

                size = os.path.getsize(file_path)
                size_on_disk = os.stat(file_path).st_size

                file_info = {
                    'File Name': os.path.basename(file_path),
                    'Type of File': os.path.splitext(file_path)[1][1:],  # Get file extension without the dot
                    'Location': os.path.abspath(file_path),
                    'Size': size,
                    'Size on Disk': size_on_disk,
                    'Created': created_time,
                    'Modified': modified_time,
                    'Accessed': accessed_time,
                    'Image Width': img.width,
                    'Image Height': img.height
                }

                return file_info
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            return None

    def process_folder(self, folder_path, column):
        all_file_info = []

        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            if os.path.isfile(file_path) and filename.lower().endswith(('.gif', '.png', '.jpg', '.jpeg')):
                file_info = self.extract_image_info(file_path)
                if file_info:
                    all_file_info.append(file_info)

        return all_file_info, column

    def apply_thick_border(self, ws, start_row, end_row, column):
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        for i in range(start_row, end_row + 1):
            cell = ws.cell(row=i, column=column)
            cell.border = thick_border
            if 'File Name' in cell.value:
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    def export_to_excel(self, folder_name, all_file_info, excel_file):
        try:
            if os.path.exists(excel_file):
                # If the Excel file already exists, load it
                wb = load_workbook(excel_file)
            else:
                # If the Excel file doesn't exist, create a new workbook
                wb = Workbook()

            ws = wb.create_sheet(title=f'Sheet_{len(wb.sheetnames) + 1}')  # Create a new sheet for the current run
            iterator = 0  # Variable for loop
            ws['A1'], ws['B1'] = "Before", "After"

            column_a = []
            column_a_1 = []
            column_a_2 = []
            column_b = []
            column_b_1 = []

            for folder_info, column in all_file_info:
                if column == 1:
                    column_a = folder_info
                else:
                    column_b = folder_info
                    
            for a in column_a:
                found = False
                for b in column_b:
                    if a['Size'] == b['Size']:
                        column_a_1.append(a)
                        column_b_1.append(b)
                        found = True
                        break
                if found == False:
                    column_a_2.append(a)

            column_b_2 = [item for item in column_b if item not in column_b_1]

            for folder_info, column in all_file_info:
                
                # ws[chr(64+column)+'1'] = folder_name[iterator]  # Name the Header
                iterator += 1
                
                start_row = len(ws[chr(64 + column)]) + 2  # Start from row 2 to leave space for headers
                
                row = 2  # Counter for row
                if column == 1:
                    for file_info in column_a_1:
                        for idx, (key, value) in enumerate(file_info.items()):
                            ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                        
                        self.apply_thick_border(ws, row, row + idx, column)
                        row = row + len(file_info)
                    for file_info in column_a_2:
                        for idx, (key, value) in enumerate(file_info.items()):
                            ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                        
                        self.apply_thick_border(ws, row, row + idx, column)
                        row = row + len(file_info)
                else:
                    for file_info in column_b_1:
                        for idx, (key, value) in enumerate(file_info.items()):
                            ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                        
                        self.apply_thick_border(ws, row, row + idx, column)
                        row = row + len(file_info)
                    for file_info in column_b_2:
                        for idx, (key, value) in enumerate(file_info.items()):
                            ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                        
                        self.apply_thick_border(ws, row, row + idx, column)
                        row = row + len(file_info)

            wb.save(excel_file)
        except Exception as e: 
            print(f"Error exporting to Excel: {e}")

    def execute_test_code(self):
        if not self.input_folder_test_code or len(self.input_folder_test_code) != 2:
            print("Please select 2 input folders first.")
            return
        
        if not self.output_folder_test_code:
            print("Please select output folder first.")
            return
        
        excel_output = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx  *.xls")])

        if excel_output:
            output_excel_file = excel_output
            folder_name = self.input_folder_test_code
            
            all_file_info_1 = self.process_folder(folder_name[0], column=1)
            all_file_info_2 = self.process_folder(folder_name[1], column=2)

            all_file_info = [all_file_info_1, all_file_info_2]
            self.export_to_excel(folder_name, all_file_info, output_excel_file)
            self.text_widget.insert(tk.END, 'Test code executed!' + '\n\n')
        else:
            print("Please select output file first.")
            return

if __name__ == "__main__":
    root = tk.Tk()
    editor = HexEditor(root)
    root.mainloop()
