from PIL import Image
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill

def extract_image_info(file_path):
    try:
        with Image.open(file_path) as img:
            created_time = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')
            modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')
            accessed_time = datetime.fromtimestamp(os.path.getatime(file_path)).strftime('%A, %B %d, %Y %I:%M:%S %p')

            size = os.path.getsize(file_path)
            size_on_disk = os.stat(file_path).st_size

            file_info = {
                'File Name': os.path.basename(file_path),
                'Type of File': os.path.splitext(file_path)[1][1:],  # Get file extension without the dot
                'Location': os.path.abspath(file_path),
                'Size': size,
                'Size on Disk': size_on_disk,
                'Created': created_time,
                'Modified': modified_time,
                'Accessed': accessed_time,
                'Image Width': img.width,
                'Image Height': img.height
            }

            return file_info
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return None

def process_folder(folder_path, column):
    all_file_info = []

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path) and filename.lower().endswith(('.gif', '.png', '.jpg', '.jpeg')):
            file_info = extract_image_info(file_path)
            if file_info:
                all_file_info.append(file_info)

    return all_file_info, column

def apply_thick_border(ws, start_row, end_row, column):
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    for i in range(start_row, end_row + 1):
        cell = ws.cell(row=i, column=column)
        cell.border = thick_border
        if 'File Name' in cell.value:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def export_to_excel(folder_name, all_file_info, excel_file):
    try:
        if os.path.exists(excel_file):
            # If the Excel file already exists, load it
            wb = load_workbook(excel_file)
        else:
            # If the Excel file doesn't exist, create a new workbook
            wb = Workbook()

        ws = wb.create_sheet(title=f'Sheet_{len(wb.sheetnames) + 1}')  # Create a new sheet for the current run
        iterator = 0  # Variable for loop
        ws['A1'], ws['B1'] = "Before", "After"

        column_a = []
        column_b = []
        for folder_info, column in all_file_info:
            
            # ws[chr(64+column)+'1'] = folder_name[iterator]  # Name the Header
            iterator += 1
            
            start_row = len(ws[chr(64 + column)]) + 2  # Start from row 2 to leave space for headers
            
            row = 2  # Counter for row
            if column == 1:
                column_a = folder_info
                for file_info in folder_info:
                    for idx, (key, value) in enumerate(file_info.items()):
                        ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                    
                    apply_thick_border(ws, row, row + idx, column)
                    row = row + len(file_info)
            else:
                column_b = folder_info
                # to re-arrange column_b
                modified_column_b = []
                for a in column_a:
                    found = False
                    for b in column_b:
                        if a['Size'] == b['Size']:
                            modified_column_b.append(b)
                            found = True
                            break
                    if found == False:
                        modified_column_b.append(
                            {
                                'File Name': '',
                                'Type of File': '',  # Get file extension without the dot
                                'Location': '',
                                'Size': '',
                                'Size on Disk': '',
                                'Created': '',
                                'Modified': '',
                                'Accessed': '',
                                'Image Width': '',
                                'Image Height': ''
                            }
                        )

                # print(modified_column_b)
                # print(len(modified_column_b))
                for file_info in modified_column_b:
                    # if file_info:
                        # print(file_info)
                    for idx, (key, value) in enumerate(file_info.items()):
                        ws.cell(row=row + idx, column=column, value=f"{key}: {value}")
                    # else:
                    #     for i in range(0,10):
                    #         ws.cell(row=row + i, column=column, value="")
                    
                    apply_thick_border(ws, row, row + idx, column)
                    row = row + len(file_info)

        wb.save(excel_file)
    except Exception as e: 
        print(f"Error exporting to Excel: {e}")

if __name__ == "__main__":
    folder_path_1 = r'xxxxxxxxxx'
    folder_path_2 = r'xxxxxxxxxx'
    output_excel_file = 'Output300File.xlsx'
    folder_name = [folder_path_1, folder_path_2]
    
    all_file_info_1 = process_folder(folder_path_1, column=1)
    all_file_info_2 = process_folder(folder_path_2, column=2)

    all_file_info = [all_file_info_1, all_file_info_2]
    export_to_excel(folder_name, all_file_info, output_excel_file)
